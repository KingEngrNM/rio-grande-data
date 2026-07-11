"""
Rio Grande USGS / Colorado DWR Data Fetcher
============================================
Retrieves streamflow discharge data for key Rio Grande gages
from Lobatos, CO to San Marcial, NM.

Data sources:
  - USGS NWIS  : all stations for daily and 15-minute discharge data
  - Colorado DWR CDSS REST API : Lobatos 15-minute telemetry only
    (USGS publishes only daily means for Lobatos; real-time telemetry
     is operated by Colorado Division of Water Resources)

Produces two Excel workbooks:
  1. RioGrande_WaterYear_Daily.xlsx
     - One sheet per station
     - Daily mean discharge from Oct 1 of current water year to present
     - Additional sheets: EB Storage & WSE, EB Release, EB Net Inflow

  2. RioGrande_30Day_15min.xlsx
     - One sheet per station
     - 15-minute instantaneous discharge for past 30 days
     - Lobatos sheet pulls from Colorado DWR telemetry API

USGS Parameter Codes:
  00060 = Discharge (cfs)

USGS Qualification Codes:
  A  = Approved   P = Provisional   e = Estimated
  Ice = Ice affected   Eqp = Equipment malfunction

Colorado DWR flagA codes (for Lobatos):
  O  = Original/observed   E = Estimated   W = Working value

Colorado DWR API notes:
  - Free account + API key recommended to avoid daily rate limits.
  - Sign up at: https://dwr.state.co.us/Rest/GET/Home/SignUp
  - Set DWR_API_KEY below once you have a key, or leave blank for
    anonymous access (1,000 calls/day, 600,000 rows/day limit).

Requirements:
  pip install requests openpyxl pandas

Usage:
  python rio_grande_usgs.py

Output files are written to the same folder as this script.
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import json
import sys
import os
import time

try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    print("NOTE: plotly not installed. HTML dashboard will be skipped.")
    print("      To enable: pip install plotly")

# =============================================================================
# OUTPUT FILES
# =============================================================================
OUTPUT_FILES = [
    "RioGrande_WaterYear_Daily.xlsx",
    "RioGrande_30Day_15min.xlsx",
    "RioGrande_Dashboard.html",
    "RioGrande_Dashboard_15min.html",
]

# When running unattended (e.g. GitHub Actions), there's no one to answer
# interactive prompts -- detect this and skip them automatically.
IS_CI = bool(os.environ.get("CI") or os.environ.get("GITHUB_ACTIONS")
              or not sys.stdin.isatty())


def safe_save(save_callable, filename):
    """
    Call save_callable() to write `filename`, handling the common case where
    the file is currently open in another program (Excel, a browser, etc.)
    and cannot be overwritten.

    On PermissionError, prompts the user to close the file and retry, or to
    skip saving this file. Returns True if the file was saved, False if it
    was skipped.

    In CI / non-interactive environments (IS_CI), a PermissionError is
    simply reported and the file is skipped -- there's no one to prompt.
    """
    while True:
        try:
            save_callable()
            print(f"Saved: {filename}")
            return True
        except PermissionError:
            print(f"\n  Could not write {filename} -- it may be open in "
                  f"another program (e.g. Excel or a browser).")
            if IS_CI:
                print(f"  Non-interactive run -- skipping {filename}.")
                return False
            choice = input("  Close the file and press Enter to retry, "
                            "or type 'skip' to skip it: ").strip().lower()
            if choice == "skip":
                print(f"  Skipped: {filename}")
                return False

# =============================================================================
# STATION CONFIGURATION
# Add or remove stations here. Order determines sheet order in workbook.
#
# Each station is a dict with:
#   usgs_site  : USGS site number (used for daily and 15-min data)
#   short_name : Short label used as the Excel sheet name (max ~28 chars)
#   full_name  : Full descriptive station name
#   dwr_abbrev : (optional) Colorado DWR station abbreviation.
#                If set, BOTH daily and 15-minute data come from Colorado
#                DWR instead of USGS. Used for Lobatos — DWR operates the
#                gage and holds the authoritative HydroBase record used for
#                Rio Grande Compact accounting.
#                Daily  : surfacewater/surfacewatertsday endpoint (approved)
#                15-min : telemetrystations/telemetrytimeseriesraw endpoint
# =============================================================================
STATIONS = [
    {
        "usgs_site":  "08251500",
        "short_name": "Lobatos",
        "full_name":  "Rio Grande near Lobatos, CO",
        "dwr_abbrev": "RIOLOBCO",   # Both daily and 15-min from Colorado DWR
    },
    {
        "usgs_site":  "08263500",
        "short_name": "Cerro",
        "full_name":  "Rio Grande near Cerro, NM",
    },
    {
        "usgs_site":  "08279500",
        "short_name": "Embudo",
        "full_name":  "Rio Grande at Embudo, NM",
    },
    {
        "usgs_site":  "08290000",
        "short_name": "Chamita",
        "full_name":  "Rio Chama near Chamita, NM",
    },
    {
        "usgs_site":  "08313000",
        "short_name": "Otowi",
        "full_name":  "Rio Grande at Otowi Bridge, NM",
    },
    {
        "usgs_site":  "08317400",
        "short_name": "Blw Cochiti",
        "full_name":  "Rio Grande below Cochiti Dam, NM",
    },
    {
        "usgs_site":  "08329918",
        "short_name": "Alameda",
        "full_name":  "Rio Grande at Alameda Bridge, NM",
    },
    {
        "usgs_site":  "08329928",
        "short_name": "Paseo del Norte",
        "full_name":  "Rio Grande near Paseo del Norte, NM",
    },
    {
        "usgs_site":  "08330000",
        "short_name": "Albuquerque",
        "full_name":  "Rio Grande at Albuquerque, NM",
    },
    {
        "usgs_site":  "08331160",
        "short_name": "Bosque Farms",
        "full_name":  "Rio Grande near Bosque Farms, NM",
    },
    {
        "usgs_site":  "08331510",
        "short_name": "Hwy 346",
        "full_name":  "Rio Grande at State Hwy 346 near Bosque, NM",
    },
    {
        "usgs_site":  "08354900",
        "short_name": "San Acacia",
        "full_name":  "Rio Grande Floodway at San Acacia, NM",
    },
    {
        "usgs_site":  "08355050",
        "short_name": "Escondida",
        "full_name":  "Rio Grande at Bridge near Escondida, NM",
    },
    {
        "usgs_site":  "08355490",
        "short_name": "Hwy 380",
        "full_name":  "Rio Grande above US Hwy 380 near San Antonio, NM",
    },
    {
        "usgs_site":  "08358300",
        "short_name": "San Marcial CC",
        "full_name":  "Rio Grande Conveyance Channel at San Marcial, NM",
    },
    {
        "usgs_site":  "08358400",
        "short_name": "San Marcial FW",
        "full_name":  "Rio Grande Floodway at San Marcial, NM",
    },
    {
        "usgs_site":  "08359500",
        "short_name": "Narrows",
        "full_name":  "Rio Grande at Narrows in Elephant Butte Res., NM",
    },
]

# =============================================================================
# COLORADO DWR API KEY (optional but recommended)
# Sign up free at: https://dwr.state.co.us/Rest/GET/Home/SignUp
# Leave as empty string "" for anonymous access.
# =============================================================================
DWR_API_KEY = ""   # e.g. "B9xxxxx-xxxx-4D47-xxxx-xxxxxxxxxxxx"

# Default USGS parameter code — discharge only
PARAM_DISCHARGE = "00060"

# =============================================================================
# RECLAMATION RISE API -- ELEPHANT BUTTE RESERVOIR
# https://data.usbr.gov/rise/api
#
# Location ID : 323  (Elephant Butte Reservoir Dam and Powerplant)
# Item / parameter IDs:
#   329  : Daily Storage-af
#   332  : Daily Elevation-ft (water surface elevation AMSL)
#   4377 : Daily Release - Total-cfs
#   4378 : Daily Release - Total-af
#
# API returns JSON:API pages; we request 500 per page and follow 'next' links.
# =============================================================================
RISE_BASE    = "https://data.usbr.gov/rise/api/result"
RISE_LOC_ID  = 323
RISE_ITEMS   = {
    "storage_af":   329,
    "elevation_ft": 332,
}
# Release comes from USGS site 08361000 (Rio Grande below Elephant Butte Dam)
EB_RELEASE_USGS_SITE = "08361000"
AF_PER_DAY_TO_CFS = 86400 / 43560   # exact unit conversion factor


def fetch_reclamation_eb(start_dt, end_dt):
    """
    Fetch daily Elephant Butte Reservoir data from the USBR RISE API.

    Retrieves storage (af) and water surface elevation (ft AMSL) for the
    date range [start_dt, end_dt] from the USBR RISE API.
    Paginates automatically using the JSON:API 'next' link.

    Arguments:
      start_dt : date object for start of period
      end_dt   : date object for end of period

    Returns a dict with keys:
      'storage_af'   : pd.Series indexed by date
      'elevation_ft' : pd.Series indexed by date
    Each series contains float values; missing days are NaN.
    Returns None if both parameters fail.
    """
    results = {}

    for param_name, item_id in RISE_ITEMS.items():
        params = {
            "catalogItemId":              item_id,
            "locationId":                 RISE_LOC_ID,
            "dateTime[after]":            start_dt.isoformat(),
            "dateTime[before]":           end_dt.isoformat(),
            "resultAttributes.timeStep":  "day",
            "itemsPerPage":               500,
            "page":                       1,
        }

        records = []
        page = 1
        while True:
            params["page"] = page
            try:
                resp = requests.get(RISE_BASE, params=params, timeout=30)
                resp.raise_for_status()
                payload = resp.json()
            except requests.exceptions.RequestException as e:
                print(f"  RISE API error for {param_name} (item {item_id}): {e}")
                break

            data_list = payload.get("data", [])
            if not data_list:
                break
            records.extend(data_list)

            # Check for another page via JSON:API links
            links = payload.get("links", {})
            if not links.get("next"):
                break
            page += 1

        if not records:
            print(f"  RISE: no data returned for {param_name} (item {item_id})")
            results[param_name] = None
            continue

        timestamps, vals = [], []
        for rec in records:
            attrs = rec.get("attributes", {})
            # Filter to only records belonging to this specific catalog item.
            # catalogItemId queries can return mixed itemIds for the same location.
            if attrs.get("itemId") != item_id:
                continue
            dt_str = attrs.get("dateTime", "")
            val    = attrs.get("result")   # confirmed key from API response
            if not dt_str or val is None:
                continue
            try:
                timestamps.append(dt_str)
                vals.append(float(val))
            except (ValueError, TypeError):
                continue

        if not timestamps:
            results[param_name] = None
            continue

        index = pd.to_datetime(timestamps, utc=True).tz_convert(None).normalize()
        s = pd.Series(vals, index=index, name=param_name).sort_index()
        # Reclamation stores multiple intraday updates under the same date;
        # after normalizing to date, keep the last (most recently updated) value.
        s = s.groupby(s.index).last()
        results[param_name] = s
        print(f"  RISE: {param_name} -- {len(s)} daily records (from {len(vals)} raw)")

    # Return None only if every parameter failed
    if all(v is None for v in results.values()):
        return None
    return results


# =============================================================================
# DATE RANGE SETUP
# =============================================================================
today = date.today()

# Water year starts October 1. If we're before Oct 1, water year started
# last calendar year.
if today.month >= 10:
    water_year_start = date(today.year, 10, 1)
else:
    water_year_start = date(today.year - 1, 10, 1)

thirty_days_ago = today - timedelta(days=30)

# =============================================================================
# USGS API FUNCTIONS (new api.waterdata.usgs.gov endpoints)
# =============================================================================
# USGS is migrating from waterservices.usgs.gov (legacy) to api.waterdata.usgs.gov.
# Legacy services showed significant degradation beginning fall 2025.
# Both daily and 15-minute functions below use the new API.
#
# New API monitoring location IDs are formatted as "USGS-08251500" (agency-number).
# Daily endpoint  : /ogcapi/v0/collections/daily/items
# 15-min endpoint : /ogcapi/v0/collections/continuous/items

USGS_NEW_BASE    = "https://api.waterdata.usgs.gov/ogcapi/v0/collections"
USGS_LEGACY_BASE = "https://waterservices.usgs.gov/nwis"

def fetch_usgs_daily(site, start_dt, end_dt):
    """
    Fetch daily mean discharge from the USGS Water Data API.

    Tries the new API (api.waterdata.usgs.gov) first. If that returns no
    data — as happens for some cooperative stations like Lobatos (08251500)
    that have not yet been migrated — falls back to the legacy API
    (waterservices.usgs.gov). The legacy API is degrading but still works
    for some stations.

    Arguments:
      site     : USGS site number string, e.g. "08251500"
      start_dt : date object for start of period
      end_dt   : date object for end of period

    Returns a dict with keys 'flow' and 'flow_cd' (pandas Series indexed
    by date), or None if both APIs fail.
    """
    # --- Try new API first ---
    url = f"{USGS_NEW_BASE}/daily/items"
    query = {
        "f":                       "json",
        "monitoring_location_id":  f"USGS-{site}",
        "parameter_code":          PARAM_DISCHARGE,
        "statistic_id":            "00003",        # daily mean
        "datetime":                f"{start_dt.isoformat()}/{end_dt.isoformat()}",
        "limit":                   50000,
        "skipGeometry":            "true",
    }

    try:
        resp = requests.get(url, params=query, timeout=30)
        resp.raise_for_status()
        data     = resp.json()
        features = data.get("features", [])

        timestamps, vals, codes = [], [], []
        for f in features:
            props  = f.get("properties", {})
            dt     = props.get("time", "")
            val    = props.get("value")
            status = props.get("approval_status", "")
            code   = "A" if status == "Approved" else "P"
            if not dt or val is None:
                continue   # skip null-value records; legacy API may have them
            timestamps.append(dt)
            vals.append(float(val))
            codes.append(code)

        if timestamps:
            index = pd.to_datetime(timestamps)
            index = index.tz_localize(None) if index.tz is None else index.tz_convert(None)
            index = index.normalize()
            series_flow = pd.Series(vals,  index=index, name="00060").sort_index()
            series_code = pd.Series(codes, index=index, name="00060_cd").sort_index()

            # If the new API returned fewer dates than expected, it has gaps
            # (null-value records were skipped). Fall through to the legacy API.
            expected_days = (end_dt - start_dt).days + 1
            if len(series_flow) >= expected_days * 0.95:   # allow 5% gap tolerance
                return {"flow": series_flow, "flow_cd": series_code}
            print(f"  New API returned only {len(series_flow)}/{expected_days} days "
                  f"for {site}, trying legacy API for complete record...")
        else:
            print(f"  New API returned no daily data for {site}, trying legacy API...")

    except requests.exceptions.RequestException as e:
        print(f"  New API error for {site}: {e}, trying legacy API...")

    # --- Fallback: legacy API ---
    url_legacy = f"{USGS_LEGACY_BASE}/dv/"
    query_legacy = {
        "format":      "json",
        "sites":       site,
        "startDT":     start_dt.isoformat(),
        "endDT":       end_dt.isoformat(),
        "parameterCd": PARAM_DISCHARGE,
        "siteStatus":  "all",
    }

    try:
        resp = requests.get(url_legacy, params=query_legacy, timeout=30)
        resp.raise_for_status()
        data       = resp.json()
        time_series = data.get("value", {}).get("timeSeries", [])

        if not time_series:
            print(f"  Legacy API also returned no daily data for {site}")
            return None

        timestamps, vals, codes = [], [], []
        for ts in time_series:
            param_code = ts["variable"]["variableCode"][0]["value"]
            if param_code != PARAM_DISCHARGE:
                continue
            for v in ts["values"][0]["value"]:
                raw = v["value"]
                if raw in ("-999999", ""):
                    continue
                timestamps.append(v["dateTime"])
                vals.append(float(raw))
                codes.append(v.get("qualifiers", ["P"])[0] if v.get("qualifiers") else "P")

        if not timestamps:
            return None

        index = pd.to_datetime(timestamps)
        index = index.tz_localize(None) if index.tz is None else index.tz_convert(None)
        index = index.normalize()
        print(f"  Legacy API succeeded for {site}")
        return {
            "flow":    pd.Series(vals,  index=index, name="00060").sort_index(),
            "flow_cd": pd.Series(codes, index=index, name="00060_cd").sort_index(),
        }

    except requests.exceptions.RequestException as e:
        print(f"  Legacy API also failed for {site}: {e}")
        return None


def fetch_usgs_continuous(site, start_dt, end_dt):
    """
    Fetch 15-minute instantaneous discharge from the new USGS Water Data API.

    Arguments:
      site     : USGS site number string, e.g. "08251500"
      start_dt : date object for start of period
      end_dt   : date object for end of period

    Returns a dict with keys 'flow' and 'flow_cd', indexed by datetime.
    Returns None if the request fails or has no data.

    New API notes:
      - endpoint: /ogcapi/v0/collections/continuous/items
      - time range: ISO interval "2026-05-01/2026-06-04"
      - same approval_status field as daily endpoint
      - up to 3 years of data per request
    """
    url = f"{USGS_NEW_BASE}/continuous/items"
    query = {
        "f":                       "json",
        "monitoring_location_id":  f"USGS-{site}",
        "parameter_code":          PARAM_DISCHARGE,
        "time":                    f"{start_dt.isoformat()}/{end_dt.isoformat()}",
        "limit":                   50000,
        "skipGeometry":            "true",
    }

    try:
        resp = requests.get(url, params=query, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"  ERROR fetching 15-min {site}: {e}")
        return None

    data     = resp.json()
    features = data.get("features", [])

    if not features:
        print(f"  No 15-min data returned for site {site}")
        return None

    timestamps = []
    vals       = []
    codes      = []

    for f in features:
        props  = f.get("properties", {})
        dt     = props.get("time", "")
        val    = props.get("value")
        status = props.get("approval_status", "")
        code   = "A" if status == "Approved" else "P"

        if not dt:
            continue
        timestamps.append(dt)
        vals.append(float(val) if val is not None else None)
        codes.append(code)

    if not timestamps:
        return None

    index = pd.to_datetime(timestamps)
    index = index.tz_localize(None) if index.tz is None else index.tz_convert(None)

    flow    = pd.Series(vals,  index=index, name="00060").sort_index()
    flow_cd = pd.Series(codes, index=index, name="00060_cd").sort_index()

    return {
        "flow":    flow,
        "flow_cd": flow_cd,
    }



# =============================================================================
# COLORADO DWR FETCH FUNCTION
# =============================================================================

DWR_BASE = "https://dwr.state.co.us/Rest/GET/api/v2"

def fetch_colorado_dwr(abbrev, start_dt, end_dt):
    """
    Fetch 15-minute telemetry data from Colorado DWR CDSS REST API.

    Arguments:
      abbrev   : DWR station abbreviation, e.g. "RIOLOBCO"
      start_dt : date object for start of period
      end_dt   : date object for end of period

    Returns a dict with keys 'flow' and 'flow_cd', each a pandas Series
    indexed by datetime. 'gage' key is None (DWR does not publish gage
    height for this station through the telemetry API).
    Returns None if the request fails.

    DWR flagA codes:
      O = Original/observed
      E = Estimated
      W = Working value (provisional)
    """
    url = f"{DWR_BASE}/telemetrystations/telemetrytimeseriesraw/"

    # Format dates as MM-DD-YYYY which DWR expects
    start_str = start_dt.strftime("%m-%d-%Y")
    end_str   = end_dt.strftime("%m-%d-%Y")

    query = {
        "format":      "json",
        "abbrev":      abbrev,
        "parameter":   "DISCHRG",
        "startDate":   start_str,
        "endDate":     end_str,
        "pageSize":    50000,
    }

    # Add API key if provided
    if DWR_API_KEY:
        query["apiKey"] = DWR_API_KEY

    try:
        resp = requests.get(url, params=query, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"  ERROR fetching DWR {abbrev}: {e}")
        return None

    data = resp.json()
    records = data.get("ResultList", [])

    if not records:
        print(f"  No DWR data returned for {abbrev}")
        return None

    timestamps = []
    vals       = []
    codes      = []

    for r in records:
        dt_str = r.get("measDateTime", "")
        val    = r.get("measValue")
        flag   = r.get("flagA", "") or ""

        if not dt_str:
            continue

        timestamps.append(dt_str)
        vals.append(float(val) if val is not None else None)
        codes.append(flag)

    if not timestamps:
        return None

    index = pd.to_datetime(timestamps, utc=True).tz_convert(None)
    flow_series = pd.Series(vals,  index=index, name="DISCHRG")
    code_series = pd.Series(codes, index=index, name="DISCHRG_cd")

    return {
        "flow":    flow_series,
        "flow_cd": code_series,
        "source":  "Colorado DWR CDSS Telemetry",
    }


def fetch_colorado_dwr_daily(abbrev, start_dt, end_dt):
    """
    Fetch approved daily mean discharge from the Colorado DWR HydroBase
    surface water time series endpoint.

    This is the authoritative source for Compact accounting — it contains
    DWR's reviewed daily means, which are the values submitted to USGS
    for cooperative publication.

    Arguments:
      abbrev   : DWR station abbreviation, e.g. "RIOLOBCO"
      start_dt : date object for start of period
      end_dt   : date object for end of period

    Returns a dict with keys 'flow' and 'flow_cd', indexed by date.
    Returns None if the request fails or returns no data.

    Response field notes:
      measDate  : ISO datetime with timezone offset, e.g. "2025-10-01T00:00:00-06:00"
      value     : daily mean discharge in cfs
      flagA     : approval status — "A" = Approved, other = provisional/working
      flagB     : accuracy rating — "Excellent", "Good", "Fair", "Poor"
      dataSource: "DWR" for HydroBase records
    """
    url = f"{DWR_BASE}/surfacewater/surfacewatertsday/"

    # DWR expects MM-DD-YYYY for this endpoint
    start_str = start_dt.strftime("%m-%d-%Y")
    end_str   = end_dt.strftime("%m-%d-%Y")

    query = {
        "format":       "json",
        "abbrev":       abbrev,
        "min-measDate": start_str,
        "max-measDate": end_str,
        "pageSize":     50000,
    }

    if DWR_API_KEY:
        query["apiKey"] = DWR_API_KEY

    try:
        resp = requests.get(url, params=query, timeout=30)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"  ERROR fetching DWR daily {abbrev}: {e}")
        return None

    data    = resp.json()
    records = data.get("ResultList", [])

    if not records:
        print(f"  No DWR daily data returned for {abbrev}")
        return None

    timestamps = []
    vals       = []
    codes      = []

    for r in records:
        dt_str = r.get("measDate", "")
        val    = r.get("value")
        flag_a = r.get("flagA", "") or ""
        flag_b = r.get("flagB", "") or ""

        if not dt_str or val is None:
            continue

        timestamps.append(dt_str)
        vals.append(float(val))
        # Combine flagA (approval) and flagB (accuracy) into one readable code
        # e.g. "A/Good", "A/Fair", or just "A" if no accuracy rating
        code = flag_a
        if flag_b:
            code = f"{flag_a}/{flag_b}"
        codes.append(code)

    if not timestamps:
        return None

    index = pd.to_datetime(timestamps, utc=True).tz_convert(None)
    index = index.normalize()   # date only, no time component

    flow    = pd.Series(vals,  index=index, name="DISCHRG").sort_index()
    flow_cd = pd.Series(codes, index=index, name="DISCHRG_cd").sort_index()

    return {
        "flow":    flow,
        "flow_cd": flow_cd,
        "source":  "Colorado DWR HydroBase (surfacewatertsday)",
    }


# =============================================================================
# EXCEL STYLING HELPERS
# =============================================================================

# Color palette
COLOR_HEADER_BG  = "1F4E79"   # Dark blue
COLOR_HEADER_FG  = "FFFFFF"   # White text
COLOR_SUBHDR_BG  = "BDD7EE"   # Light blue
COLOR_APPROVED   = "E2EFDA"   # Light green - approved data
COLOR_PROVISIONAL= "FFF2CC"   # Light yellow - provisional data
COLOR_ALT_ROW    = "F5F5F5"   # Light grey alternating row

THIN_BORDER = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin"),
)

def style_header(cell, text, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, size=11):
    """Apply bold header styling to a cell."""
    cell.value = text
    cell.font  = Font(name="Arial", bold=True, color=fg, size=size)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def style_data(cell, value, number_format=None, bg=None):
    """Apply data cell styling."""
    cell.value = value
    cell.font  = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")
    cell.border = THIN_BORDER
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if number_format:
        cell.number_format = number_format

def autofit_columns(ws, min_width=10, max_width=40):
    """Set reasonable column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# =============================================================================
# WORKBOOK BUILDERS
# =============================================================================

def build_daily_sheet(ws, station, data):
    """
    Write daily mean data to a worksheet.

    Layout:
      Row 1: Station title (merged)
      Row 2: Data source note
      Row 3: Blank
      Row 4: Column headers
      Row 5+: Data rows (one per day)
      Then: blank row, field visits table if available
    """
    site_no    = station["usgs_site"]
    short_name = station["short_name"]
    full_name  = station["full_name"]

    # --- Title block ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    style_header(title_cell,
                 f"{full_name}   (USGS {site_no})",
                 bg=COLOR_HEADER_BG, size=12)

    ws.merge_cells("A2:E2")
    note = ws["A2"]
    note.value = (f"Daily mean discharge | Water year {water_year_start.year} "
                  f"({water_year_start} to {today}) | "
                  f"Retrieved {date.today()}")
    note.font  = Font(name="Arial", italic=True, size=9, color="595959")
    note.alignment = Alignment(horizontal="left")

    # --- Column headers (row 4) ---
    headers = ["Date", "Discharge (cfs)", "Qual. Code"]
    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=col), hdr,
                     bg=COLOR_SUBHDR_BG, fg="000000")

    # --- Data rows ---
    current_row = 5
    if data is None or data.get("flow") is None:
        ws.cell(row=current_row, column=1).value = "No data available for this period."
    else:
        flow_series = data["flow"]
        code_series = data.get("flow_cd")

        for dt, flow_val in flow_series.items():
            date_str = dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt)
            code     = code_series[dt] if (code_series is not None and dt in code_series.index) else ""

            # Choose row background based on approval code
            if "A" in str(code):
                row_bg = COLOR_APPROVED
            elif "P" in str(code):
                row_bg = COLOR_PROVISIONAL
            else:
                row_bg = None

            style_data(ws.cell(row=current_row, column=1), date_str, bg=row_bg)
            style_data(ws.cell(row=current_row, column=2),
                       round(flow_val, 1) if flow_val is not None else "M",
                       number_format="#,##0.0", bg=row_bg)
            style_data(ws.cell(row=current_row, column=3), code, bg=row_bg)
            current_row += 1

    # --- Legend ---
    current_row += 1
    ws.cell(row=current_row, column=1).value = (
        "USGS Qual. Codes:  A = Approved   P = Provisional   "
        "e = Estimated   Ice = Ice affected   Eqp = Equipment malfunction   "
        "M = Missing   ||   "
        "Colorado DWR (Lobatos):  A/Good, A/Fair, A/Poor = Approved + accuracy rating"
    )
    ws.cell(row=current_row, column=1).font = Font(name="Arial", italic=True, size=8, color="595959")
    ws.merge_cells(f"A{current_row}:C{current_row}")

    autofit_columns(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A5"   # Keep headers visible when scrolling


def build_15min_sheet(ws, station, data):
    """
    Write 15-minute instantaneous discharge data to a worksheet.
    Handles both USGS and Colorado DWR data formats.
    """
    site_no    = station["usgs_site"]
    short_name = station["short_name"]
    full_name  = station["full_name"]
    dwr_abbrev = station.get("dwr_abbrev")

    # Determine data source label for the header note
    if dwr_abbrev and data is not None and data.get("source"):
        source_label = data["source"]
    elif dwr_abbrev:
        source_label = "Colorado DWR CDSS Telemetry"
    else:
        source_label = f"USGS {site_no}"

    ws.merge_cells("A1:C1")
    style_header(ws["A1"],
                 f"{full_name}   (USGS {site_no})",
                 bg=COLOR_HEADER_BG, size=12)

    ws.merge_cells("A2:C2")
    note = ws["A2"]
    note.value = (f"15-minute instantaneous discharge | "
                  f"{thirty_days_ago} to {today} (past 30 days) | "
                  f"Source: {source_label} | "
                  f"Retrieved {date.today()}")
    note.font  = Font(name="Arial", italic=True, size=9, color="595959")
    note.alignment = Alignment(horizontal="left")

    headers = ["Date/Time", "Discharge (cfs)", "Qual. Code"]
    for col, hdr in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=col), hdr,
                     bg=COLOR_SUBHDR_BG, fg="000000")

    current_row = 5
    if data is None or data.get("flow") is None:
        ws.cell(row=current_row, column=1).value = "No data available for this period."
    else:
        flow_series = data["flow"]
        code_series = data.get("flow_cd")

        for dt, flow_val in flow_series.items():
            dt_str = dt.strftime("%Y-%m-%d %H:%M") if hasattr(dt, "strftime") else str(dt)
            code   = code_series[dt] if (code_series is not None and dt in code_series.index) else ""

            if "A" in str(code):
                row_bg = COLOR_APPROVED
            elif "P" in str(code):
                row_bg = COLOR_PROVISIONAL
            else:
                row_bg = None

            style_data(ws.cell(row=current_row, column=1), dt_str, bg=row_bg)
            style_data(ws.cell(row=current_row, column=2),
                       round(flow_val, 1) if flow_val is not None else "M",
                       number_format="#,##0.0", bg=row_bg)
            style_data(ws.cell(row=current_row, column=3), code, bg=row_bg)
            current_row += 1

    # --- Legend ---
    current_row += 1
    ws.cell(row=current_row, column=1).value = (
        "USGS Qual. Codes:  A = Approved   P = Provisional   "
        "e = Estimated   Ice = Ice affected   Eqp = Equipment malfunction   "
        "M = Missing   ||   "
        "Colorado DWR Flag Codes (Lobatos):  O = Original/observed   "
        "E = Estimated   W = Working/provisional"
    )
    ws.cell(row=current_row, column=1).font = Font(name="Arial", italic=True, size=8, color="595959")
    ws.merge_cells(f"A{current_row}:C{current_row}")

    autofit_columns(ws)
    ws.column_dimensions["A"].width = 18   # wider for datetime
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A5"


def add_cover_sheet(wb, title, date_range_str, stations, eb_sheets=None):
    """Add a cover/index sheet as the first sheet.

    eb_sheets (optional): list of (short_name, description) tuples for
    Elephant Butte tabs that are appended below the station list.
    """
    ws = wb.create_sheet(title="Index", index=0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    style_header(ws["A1"], title, bg=COLOR_HEADER_BG, size=14)

    ws.merge_cells("A2:D2")
    ws["A2"].value = date_range_str
    ws["A2"].font  = Font(name="Arial", italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1).value = f"Generated: {date.today()}"
    ws.cell(row=3, column=1).font  = Font(name="Arial", size=9, color="595959")

    ws.cell(row=5, column=1).value = "Station"
    ws.cell(row=5, column=2).value = "USGS Site No."
    ws.cell(row=5, column=3).value = "Full Name"
    for col in range(1, 4):
        style_header(ws.cell(row=5, column=col), ws.cell(row=5, column=col).value,
                     bg=COLOR_SUBHDR_BG, fg="000000")

    for i, station in enumerate(stations, start=6):
        ws.cell(row=i, column=1).value = station["short_name"]
        ws.cell(row=i, column=2).value = station["usgs_site"]
        ws.cell(row=i, column=3).value = station["full_name"]
        for col in range(1, 4):
            ws.cell(row=i, column=col).font   = Font(name="Arial", size=10)
            ws.cell(row=i, column=col).border = THIN_BORDER
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = PatternFill("solid", start_color=COLOR_ALT_ROW)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50

    # Optional Elephant Butte section
    eb_row_start = 6 + len(stations) + 1   # one blank row after last station
    if eb_sheets:
        ws.cell(row=eb_row_start, column=1).value = "Elephant Butte Reservoir"
        ws.cell(row=eb_row_start, column=1).font  = Font(name="Arial", bold=True, size=10)
        for j, (eb_name, eb_desc) in enumerate(eb_sheets, start=eb_row_start + 1):
            ws.cell(row=j, column=1).value = eb_name
            ws.cell(row=j, column=2).value = "USBR RISE"
            ws.cell(row=j, column=3).value = eb_desc
            for col in range(1, 4):
                ws.cell(row=j, column=col).font   = Font(name="Arial", size=10)
                ws.cell(row=j, column=col).border = THIN_BORDER
                if j % 2 == 0:
                    ws.cell(row=j, column=col).fill = PatternFill("solid", start_color=COLOR_ALT_ROW)
        footnote_offset = eb_row_start + 1 + len(eb_sheets) + 1
    else:
        footnote_offset = eb_row_start + 1

    # Footnote goes two rows below the last station row (stations start at row 6)
    footnote_row = footnote_offset
    footnote_ref = f"A{footnote_row}:D{footnote_row}"
    ws.merge_cells(footnote_ref)
    ws[f"A{footnote_row}"].value = (
        "Data source: U.S. Geological Survey National Water Information System (NWIS) | "
        "waterservices.usgs.gov  |  Data are provisional and subject to revision."
    )
    ws[f"A{footnote_row}"].font      = Font(name="Arial", italic=True, size=8, color="595959")
    ws[f"A{footnote_row}"].alignment = Alignment(horizontal="left", wrap_text=True)



# =============================================================================
# ELEPHANT BUTTE RESERVOIR SHEET BUILDERS
# =============================================================================

def build_eb_storage_sheet(ws, eb_data):
    """
    Write daily Elephant Butte storage and water surface elevation to a sheet.
    Data source: USBR RISE API.  Not plotted in the dashboard.

    Columns: Date | Storage (af) | WSE (ft AMSL)
    """
    ws.merge_cells("A1:C1")
    style_header(ws["A1"],
                 "Elephant Butte Reservoir -- Storage and Water Surface Elevation",
                 bg=COLOR_HEADER_BG, size=12)

    ws.merge_cells("A2:C2")
    ws["A2"].value = (f"Water year {water_year_start.year} "
                      f"({water_year_start} to {today}) | "
                      f"Source: USGS site {EB_RELEASE_USGS_SITE} (waterdata.usgs.gov) | "
                      "Provisional data subject to revision")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    headers = ["Date", "Storage (acre-feet)", "Water Surface Elev. (ft AMSL)"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=col), h, bg=COLOR_SUBHDR_BG, fg="000000")

    storage   = eb_data.get("storage_af")   if eb_data else None
    elevation = eb_data.get("elevation_ft") if eb_data else None

    # Build a unified date index spanning both series
    all_dates = set()
    if storage   is not None: all_dates.update(storage.index)
    if elevation is not None: all_dates.update(elevation.index)

    for row_idx, dt in enumerate(sorted(all_dates), start=5):
        s_val = float(storage.loc[dt])   if (storage   is not None and dt in storage.index)   else float("nan")
        e_val = float(elevation.loc[dt]) if (elevation is not None and dt in elevation.index) else float("nan")
        bg = COLOR_ALT_ROW if row_idx % 2 == 0 else None

        style_data(ws.cell(row=row_idx, column=1), dt.date() if hasattr(dt, "date") else dt, bg=bg)
        ws.cell(row=row_idx, column=1).number_format = "YYYY-MM-DD"

        style_data(ws.cell(row=row_idx, column=2),
                   None if (s_val is None or s_val != s_val) else s_val,
                   number_format="#,##0", bg=bg)
        style_data(ws.cell(row=row_idx, column=3),
                   None if (e_val is None or e_val != e_val) else e_val,
                   number_format="#,##0.00", bg=bg)

    autofit_columns(ws)


def build_eb_release_sheet(ws, eb_data):
    """
    Write daily Elephant Butte release (cfs and af/day) to a sheet.
    Data source: USBR RISE API.

    Columns: Date | Release (cfs) | Release (af/day)
    """
    ws.merge_cells("A1:C1")
    style_header(ws["A1"],
                 "Elephant Butte Reservoir -- Daily Release",
                 bg=COLOR_HEADER_BG, size=12)

    ws.merge_cells("A2:C2")
    ws["A2"].value = (f"Water year {water_year_start.year} "
                      f"({water_year_start} to {today}) | "
                      "Source: USBR RISE API (data.usbr.gov) | "
                      "Provisional data subject to revision")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    headers = ["Date", "Release (cfs)", "Release (acre-feet/day)"]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=col), h, bg=COLOR_SUBHDR_BG, fg="000000")

    rel_cfs = eb_data.get("release_cfs") if eb_data else None
    rel_af  = eb_data.get("release_af")  if eb_data else None

    all_dates = set()
    if rel_cfs is not None: all_dates.update(rel_cfs.index)
    if rel_af  is not None: all_dates.update(rel_af.index)

    for row_idx, dt in enumerate(sorted(all_dates), start=5):
        cfs_val = float(rel_cfs.loc[dt]) if (rel_cfs is not None and dt in rel_cfs.index) else float("nan")
        af_val  = float(rel_af.loc[dt])  if (rel_af  is not None and dt in rel_af.index)  else float("nan")
        bg = COLOR_ALT_ROW if row_idx % 2 == 0 else None

        style_data(ws.cell(row=row_idx, column=1), dt.date() if hasattr(dt, "date") else dt, bg=bg)
        ws.cell(row=row_idx, column=1).number_format = "YYYY-MM-DD"

        style_data(ws.cell(row=row_idx, column=2),
                   None if (cfs_val is None or cfs_val != cfs_val) else cfs_val,
                   number_format="#,##0.000000", bg=bg)
        style_data(ws.cell(row=row_idx, column=3),
                   None if (af_val is None or af_val != af_val) else af_val,
                   number_format="#,##0.000000", bg=bg)

    autofit_columns(ws)


def build_eb_netinflow_sheet(ws, eb_data):
    """
    Write daily Elephant Butte net inflow (af/day and cfs) to a sheet,
    and return two pd.Series for dashboard use.

    Net inflow (Rio Grande Compact definition):
      delta_storage_af  = S(t) - S(t-1)       [acre-feet/day]
      net_inflow_af     = release_af + delta_storage_af
      net_inflow_cfs    = net_inflow_af * (86400 / 43560)

    Columns: Date | Delta Storage (af/day) | Release (af/day)
                  | Net Inflow (af/day)    | Net Inflow (cfs)

    Returns (net_inflow_cfs_series, release_cfs_series) for dashboard,
    or (None, None) if data unavailable.
    """
    ws.merge_cells("A1:E1")
    style_header(ws["A1"],
                 "Elephant Butte Reservoir -- Net Inflow (Rio Grande Compact Definition)",
                 bg=COLOR_HEADER_BG, size=12)

    ws.merge_cells("A2:E2")
    ws["A2"].value = (f"Water year {water_year_start.year} "
                      f"({water_year_start} to {today}) | "
                      "Net Inflow = Release (af) + Delta Storage (af) | "
                      "Delta Storage = S(t) - S(t-1) | "
                      "cfs = af/day * (86400/43560) | "
                      f"Release source: USGS site {EB_RELEASE_USGS_SITE} | "
                      "Storage source: USBR RISE API | Provisional data subject to revision")
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", wrap_text=True)

    headers = [
        "Date",
        "Delta Storage (af/day)  S(t)-S(t-1)",
        "Release (af/day)",
        "Net Inflow (af/day)",
        "Net Inflow (cfs)",
    ]
    for col, h in enumerate(headers, start=1):
        style_header(ws.cell(row=4, column=col), h, bg=COLOR_SUBHDR_BG, fg="000000")

    if eb_data is None:
        return None, None

    storage  = eb_data.get("storage_af")
    rel_af   = eb_data.get("release_af")
    rel_cfs  = eb_data.get("release_cfs")

    if storage is None or rel_af is None:
        return None, rel_cfs

    # Compute delta storage: S(t) - S(t-1) on the storage series sorted by date
    storage_sorted  = storage.sort_index()
    delta_storage   = storage_sorted.diff()   # NaN on first day (no prior value)

    # Drive date index from storage (complete daily series) so every date
    # appears as a row even when release data is missing for that day.
    all_dates = sorted(delta_storage.index)

    ni_af_vals  = []
    ni_cfs_vals = []
    ni_af_index = []

    for row_idx, dt in enumerate(all_dates, start=5):
        ds_val  = delta_storage.loc[dt] if dt in delta_storage.index else None
        raf_val = rel_af.loc[dt]       if dt in rel_af.index       else None
        bg = COLOR_ALT_ROW if row_idx % 2 == 0 else None

        # Net inflow -- only defined when both components are present and finite
        import math
        ds_ok  = ds_val  is not None and not math.isnan(float(ds_val))
        raf_ok = raf_val is not None and not math.isnan(float(raf_val))

        if ds_ok and raf_ok:
            ni_af  = float(raf_val) + float(ds_val)
            ni_cfs = ni_af / AF_PER_DAY_TO_CFS   # af/day -> cfs
            ni_af_vals.append(ni_af)
            ni_cfs_vals.append(ni_cfs)
            ni_af_index.append(dt)
        else:
            ni_af  = None
            ni_cfs = None

        style_data(ws.cell(row=row_idx, column=1), dt.date() if hasattr(dt, "date") else dt, bg=bg)
        ws.cell(row=row_idx, column=1).number_format = "YYYY-MM-DD"

        style_data(ws.cell(row=row_idx, column=2),
                   float(ds_val) if ds_ok else None,
                   number_format="#,##0.000000", bg=bg)
        style_data(ws.cell(row=row_idx, column=3),
                   float(raf_val) if raf_ok else None,
                   number_format="#,##0.000000", bg=bg)
        style_data(ws.cell(row=row_idx, column=4),
                   ni_af, number_format="#,##0.000000", bg=bg)
        style_data(ws.cell(row=row_idx, column=5),
                   ni_cfs, number_format="#,##0.000000", bg=bg)

    autofit_columns(ws)

    ni_cfs_series = pd.Series(ni_cfs_vals, index=ni_af_index, name="eb_net_inflow_cfs")
    return ni_cfs_series, rel_cfs


# =============================================================================
# DASHBOARD CONFIGURATION
# Controls which stations appear in the HTML hydrograph and which are
# visible by default. All stations in this list must have a usgs_site
# defined in STATIONS above. Add new gages here freely — they just need
# a USGS site number, label, and whether they're on by default.
#
# default_visible = True  : line shows on load (legend entry is bold)
# default_visible = False : line is hidden on load but toggleable in legend
# =============================================================================
DASHBOARD_STATIONS = [
    # site_no       label                                   default_visible  color          dash        width
    ("08251500",   "Lobatos, CO",                           True,            "#7B2D8B",     "solid",    3),
    ("08263500",   "Cerro, NM",                             False,           "#7B2D8B",     "solid",    1.5),
    ("08279500",   "Embudo, NM",                            False,           "#7B2D8B",     "dash",     1.5),
    ("08290000",   "Chamita (Rio Chama), NM",               False,           "#7B2D8B",     "dot",      1.5),
    ("08313000",   "Otowi Bridge, NM",                      True,            "#1f77b4",     "solid",    3),
    ("08317400",   "Below Cochiti Dam, NM",                 False,           "#1f77b4",     "solid",    1.5),
    ("08329918",   "Alameda Bridge, NM",                    False,           "#2ca02c",     "dash",     3),
    ("08329928",   "Paseo del Norte, NM",                   False,           "#2ca02c",     "solid",    1.5),
    ("08330000",   "Albuquerque, NM",                       True,            "#2ca02c",     "solid",    1.5),
    ("08331160",   "Bosque Farms, NM",                      False,           "#DAA520",     "solid",    3),
    ("08331510",   "Hwy 346 near Bosque, NM",               False,           "#DAA520",     "solid",    1.5),
    ("08354900",   "San Acacia Floodway, NM",               False,           "#FF7F0E",     "solid",    3),
    ("08355050",   "Escondida, NM",                         False,           "#FF7F0E",     "solid",    1.5),
    ("08355490",   "Hwy 380 near San Antonio, NM",          False,           "#FF7F0E",     "dash",     1.5),
    ("08358300",   "San Marcial Conv. Channel, NM",         False,           "#d62728",     "solid",    1.5),
    ("08358400",   "San Marcial Floodway, NM",              True,            "#d62728",     "solid",    3),
    ("08359500",   "Narrows (Elephant Butte), NM",          False,           "#d62728",     "dash",     1.5),
    # Elephant Butte synthetic traces (keyed by string, not USGS site)
    ("EB_RELEASE",   "EB Release (cfs)",                    True,            "#000000",     "solid",    2),
    ("EB_NETINFLOW", "EB Net Inflow (cfs)",                 True,            "#000000",     "dot",      1.5),
]


def build_dashboard(station_data_map):
    """
    Build an interactive Plotly HTML hydrograph for the Rio Grande dashboard.

    Arguments:
      station_data_map : dict mapping usgs_site -> data dict (from fetch functions).
                         Built in main() as data is fetched.

    Output:
      Writes RioGrande_Dashboard.html to the current directory.

    Design:
      - Single chart, time on x-axis, discharge (cfs) on y-axis
      - Water year daily means as the primary trace
      - Legend entries are clickable to show/hide individual stations
      - Default stations visible on load; others hidden but toggleable
      - Logarithmic y-axis option commented out below — uncomment if the
        low-flow stations (Lobatos in dry years) get swamped by Otowi peaks
    """
    if not PLOTLY_AVAILABLE:
        return

    fig = go.Figure()

    # Extract our color list in station order so Plotly's auto-colorway
    # never overrides our explicit line colors
    colorway = [s[3] for s in DASHBOARD_STATIONS]
    fig.update_layout(colorway=colorway)

    for site_no, label, visible, color, dash, width in DASHBOARD_STATIONS:
        data = station_data_map.get(site_no)
        flow = data["flow"].dropna() if (data and data.get("flow") is not None) else None

        fig.add_trace(go.Scatter(
            x=flow.index if flow is not None else [],
            y=flow.values if flow is not None else [],
            name=label,
            visible=True if visible else "legendonly",
            line=dict(color=color, width=width, dash=dash),
            mode="lines",
            hovertemplate="%{fullData.name}<br>%{x|%Y-%m-%d}: %{y:,.0f} cfs<extra></extra>",
        ))

    fig.update_layout(
        title=dict(
            text=(f"Rio Grande Daily Discharge — Water Year {water_year_start.year}"
                  f"  <span style='font-size:13px; color:#888'>"
                  f"(Oct 1, {water_year_start.year} – {today})</span>"),
            font=dict(size=18),
            x=0.5,
            xanchor="center",
        ),
        xaxis=dict(
            title="Date",
            showgrid=True,
            gridcolor="#e0e0e0",
            tickformat="%b %d",
            rangeslider=dict(visible=True),   # scroll bar below chart
        ),
        yaxis=dict(
            title="Discharge (cfs)",
            showgrid=True,
            gridcolor="#e0e0e0",
            rangemode="normal",
            zeroline=True,
            zerolinecolor="#000000",
            zerolinewidth=1.5,
            # Uncomment the next two lines for log scale:
            # type="log",
            # title="Discharge (cfs, log scale)",
        ),
        legend=dict(
            title=dict(text="Click to show/hide  ▼"),
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="#cccccc",
            borderwidth=1,
            itemclick="toggle",
            itemdoubleclick="toggleothers",  # double-click isolates one station
        ),
        hovermode="closest",
        plot_bgcolor="white",
        paper_bgcolor="white",
        height=600,
        margin=dict(l=70, r=30, t=80, b=80),
        annotations=[dict(
            text=(f"Sources: USGS NWIS / Colorado DWR HydroBase  |  "
                  f"Generated {today}  |  "
                  f"Provisional data subject to revision"),
            xref="paper", yref="paper",
            x=0.0, y=-0.12,
            showarrow=False,
            font=dict(size=10, color="#888"),
            align="left",
        )],
    )

    # Write self-contained HTML — no internet connection needed to view
    html_file = "RioGrande_Dashboard.html"
    safe_save(lambda: fig.write_html(
        html_file,
        include_plotlyjs=True,    # embeds plotly.js (~3MB) so file is standalone
        full_html=True,
        config={
            "displayModeBar": True,
            "modeBarButtonsToRemove": ["lasso2d", "select2d"],
            "toImageButtonOptions": {
                "format": "png",
                "filename": f"RioGrande_WY{water_year_start.year}",
                "height": 600,
                "width": 1200,
                "scale": 2,
            },
        },
    ), html_file)


def build_15min_dashboard(station_data_map):
    """
    Build an interactive Plotly HTML hydrograph for 15-minute discharge data.

    Arguments:
      station_data_map : dict mapping usgs_site -> data dict (from fetch functions).
                         Built in main() as 15-min data is fetched.

    Output:
      Writes RioGrande_Dashboard_15min.html to the current directory.

    Design mirrors build_dashboard() exactly — same stations, colors, line
    styles, and default visibility — but plots 15-minute instantaneous values
    over the past 30 days instead of daily means over the water year.
    Hover template shows datetime to the minute.
    """
    if not PLOTLY_AVAILABLE:
        return

    fig = go.Figure()

    colorway = [s[3] for s in DASHBOARD_STATIONS]
    fig.update_layout(colorway=colorway)

    for site_no, label, visible, color, dash, width in DASHBOARD_STATIONS:
        data = station_data_map.get(site_no)
        flow = data["flow"].dropna() if (data and data.get("flow") is not None) else None

        fig.add_trace(go.Scatter(
            x=flow.index if flow is not None else [],
            y=flow.values if flow is not None else [],
            name=label,
            visible=True if visible else "legendonly",
            line=dict(color=color, width=width, dash=dash),
            mode="lines",
            hovertemplate="%{fullData.name}<br>%{x|%Y-%m-%d %H:%M}: %{y:,.0f} cfs<extra></extra>",
        ))

    fig.update_layout(
        title=dict(
            text=(f"Rio Grande 15-Minute Discharge — Past 30 Days"
                  f"  <span style='font-size:13px; color:#888'>"
                  f"({thirty_days_ago} – {today})</span>"),
            font=dict(size=18),
            x=0.5,
            xanchor="center",
        ),
        xaxis=dict(
            title="Date",
            showgrid=True,
            gridcolor="#e0e0e0",
            tickformat="%b %d",
            rangeslider=dict(visible=True),
        ),
        yaxis=dict(
            title="Discharge (cfs)",
            showgrid=True,
            gridcolor="#e0e0e0",
            rangemode="normal",
            zeroline=True,
            zerolinecolor="#000000",
            zerolinewidth=1.5,
            # Uncomment the next two lines for log scale:
            # type="log",
            # title="Discharge (cfs, log scale)",
        ),
        legend=dict(
            title=dict(text="Click to show/hide  ▼"),
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="#cccccc",
            borderwidth=1,
            itemclick="toggle",
            itemdoubleclick="toggleothers",
        ),
        hovermode="closest",
        plot_bgcolor="white",
        paper_bgcolor="white",
        height=600,
        margin=dict(l=70, r=30, t=80, b=80),
        annotations=[dict(
            text=(f"Sources: USGS NWIS / Colorado DWR HydroBase  |  "
                  f"Generated {today}  |  "
                  f"Provisional data subject to revision"),
            xref="paper", yref="paper",
            x=0.0, y=-0.12,
            showarrow=False,
            font=dict(size=10, color="#888"),
            align="left",
        )],
    )

    plot_html = fig.to_html(
        full_html=False,
        include_plotlyjs=True,
        div_id="rg15min",
        config={
            "displayModeBar": True,
            "modeBarButtonsToRemove": ["lasso2d", "select2d"],
            "toImageButtonOptions": {
                "format": "png",
                "filename": f"RioGrande_15min_{thirty_days_ago}_to_{today}",
                "height": 600,
                "width": 1200,
                "scale": 2,
            },
        },
    )

    controls_html = """
    <div style="max-width:1200px; margin: 8px auto 0; padding: 8px 12px;
                font-family: Arial, sans-serif; font-size: 13px; color: #333;
                background: #f7f7f7; border: 1px solid #ddd; border-radius: 4px;">
      <strong>Y-axis range:</strong>
      &nbsp; Min <input type="number" id="yaxis-min" style="width:90px;" placeholder="auto">
      &nbsp; Max <input type="number" id="yaxis-max" style="width:90px;" placeholder="auto">
      &nbsp;
      <button onclick="applyYRange()">Apply</button>
      <button onclick="resetYRange()">Reset (auto)</button>
      <span style="color:#888;"> &mdash; enter cfs values, e.g. 0 and 2000</span>
    </div>
    <script>
      function applyYRange() {
        var minVal = document.getElementById('yaxis-min').value;
        var maxVal = document.getElementById('yaxis-max').value;
        if (minVal === '' || maxVal === '') {
          alert('Enter both a minimum and maximum value.');
          return;
        }
        Plotly.relayout('rg15min', {
          'yaxis.range':     [parseFloat(minVal), parseFloat(maxVal)],
          'yaxis.autorange': false
        });
      }
      function resetYRange() {
        document.getElementById('yaxis-min').value = '';
        document.getElementById('yaxis-max').value = '';
        Plotly.relayout('rg15min', {'yaxis.autorange': true});
      }
    </script>
    """

    html_doc = f"""<html>
<head><meta charset="utf-8" /><title>Rio Grande 15-Minute Discharge</title></head>
<body>
{controls_html}
{plot_html}
</body>
</html>"""

    html_file = "RioGrande_Dashboard_15min.html"

    def _write_15min_html():
        with open(html_file, "w", encoding="utf-8") as f:
            f.write(html_doc)

    safe_save(_write_15min_html, html_file)


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("Rio Grande USGS Data Fetcher")
    print("=" * 60)
    print(f"Water year start : {water_year_start}")
    print(f"30-day start     : {thirty_days_ago}")
    print(f"End date         : {today}")
    print(f"Stations         : {len(STATIONS)}")
    print()

    # --- Check for existing output files before doing any work ---
    existing = [f for f in OUTPUT_FILES if os.path.exists(f)]
    if existing and not IS_CI:
        print("The following output files already exist in this folder:")
        for f in existing:
            print(f"  - {f}")
        answer = input("Overwrite them when finished? [Y/n]: ").strip().lower()
        if answer in ("n", "no"):
            print("\nAborted -- no files were changed.")
            return
        print()
    elif existing:
        print("Existing output files will be overwritten (non-interactive run).")
        print()

    # --- Workbook 1: Daily water year ---
    wb_daily = Workbook()
    wb_daily.remove(wb_daily.active)
    EB_INDEX_SHEETS = [
        ("EB Storage & WSE",  "Elephant Butte daily storage (af) and water surface elevation (ft AMSL)"),
        ("EB Release",        "Elephant Butte daily release: cfs and acre-feet/day"),
        ("EB Net Inflow",     "Calculated net inflow: Release(af) + DeltaStorage(af); also in cfs"),
    ]
    add_cover_sheet(
        wb_daily,
        "Rio Grande Daily Data — Current Water Year",
        f"Oct 1, {water_year_start.year} through {today}",
        STATIONS,
        eb_sheets=EB_INDEX_SHEETS,
    )

    # --- Workbook 2: 15-minute, 30 days ---
    wb_15min = Workbook()
    wb_15min.remove(wb_15min.active)
    add_cover_sheet(
        wb_15min,
        "Rio Grande 15-Minute Data — Past 30 Days",
        f"{thirty_days_ago} through {today}",
        STATIONS,
    )

    # Collect daily and 15-min data for dashboards as we fetch it
    dashboard_data       = {}   # site_no -> daily data dict
    dashboard_data_15min = {}   # site_no -> 15-min data dict

    # --- Fetch and write data for each station ---
    for station in STATIONS:
        site_no    = station["usgs_site"]
        short_name = station["short_name"]
        dwr_abbrev = station.get("dwr_abbrev")

        print(f"Fetching: {short_name} ({site_no})")

        # -- Daily data --
        if dwr_abbrev:
            print(f"  Daily data from Colorado DWR HydroBase ({dwr_abbrev}, "
                  f"{water_year_start} to {today})...")
            daily_data = fetch_colorado_dwr_daily(dwr_abbrev, water_year_start, today)
        else:
            print(f"  Daily data ({water_year_start} to {today})...")
            daily_data = fetch_usgs_daily(site_no, water_year_start, today)

        # Store for dashboard
        dashboard_data[site_no] = daily_data

        # Write daily sheet
        ws_d = wb_daily.create_sheet(title=short_name)
        build_daily_sheet(ws_d, station, daily_data)

        # -- 15-minute data --
        if dwr_abbrev:
            print(f"  15-min data from Colorado DWR ({dwr_abbrev}, "
                  f"{thirty_days_ago} to {today})...")
            iv_data = fetch_colorado_dwr(dwr_abbrev, thirty_days_ago, today)
        else:
            print(f"  15-min data from USGS ({thirty_days_ago} to {today})...")
            iv_data = fetch_usgs_continuous(site_no, thirty_days_ago, today)

        # Write 15-min sheet
        ws_i = wb_15min.create_sheet(title=short_name)
        build_15min_sheet(ws_i, station, iv_data)

        # Store for 15-min dashboard
        dashboard_data_15min[site_no] = iv_data

        print(f"  Done.")
        time.sleep(0.5)   # brief pause to avoid USGS API rate limits

    # --- Fetch Elephant Butte Reservoir data (daily only) ---
    print("\nFetching: Elephant Butte Reservoir (USBR RISE API)")
    print(f"  Storage and elevation ({water_year_start} to {today})...")
    eb_data = fetch_reclamation_eb(water_year_start, today)
    if eb_data is None:
        eb_data = {}

    # Release from USGS 08361000 (Rio Grande below Elephant Butte Dam)
    print(f"  Release from USGS {EB_RELEASE_USGS_SITE} ({water_year_start} to {today})...")
    release_daily = fetch_usgs_daily(EB_RELEASE_USGS_SITE, water_year_start, today)
    if release_daily and release_daily.get("flow") is not None:
        rel_cfs = release_daily["flow"]

        # Check for missing dates and fill gaps from 15-min data if possible.
        # USGS sometimes has a lag in computing daily means even when 15-min
        # data is available (e.g. around station visits or data reviews).
        full_range  = pd.date_range(water_year_start, today, freq="D")
        missing_days = full_range.difference(rel_cfs.dropna().index)
        if len(missing_days) > 0:
            print(f"  Daily record missing {len(missing_days)} date(s): "
                  f"{[str(d.date()) for d in missing_days]} -- trying 15-min fallback...")
            iv_start = missing_days.min().date()
            iv_end   = missing_days.max().date()
            iv_data  = fetch_usgs_continuous(EB_RELEASE_USGS_SITE, iv_start, iv_end)
            if iv_data and iv_data.get("flow") is not None:
                iv_flow = iv_data["flow"].dropna()
                # Resample 15-min to daily mean, keep only the missing dates
                iv_daily = iv_flow.resample("D").mean()
                iv_daily = iv_daily[iv_daily.index.isin(missing_days)]
                if len(iv_daily):
                    rel_cfs = pd.concat([rel_cfs, iv_daily]).sort_index()
                    rel_cfs = rel_cfs.groupby(rel_cfs.index).first()  # prefer daily mean
                    print(f"  Filled {len(iv_daily)} missing date(s) from 15-min data "
                          f"(computed daily mean).")

        rel_af = rel_cfs * AF_PER_DAY_TO_CFS   # cfs -> af/day
        rel_af.name = "release_af"
        eb_data["release_cfs"] = rel_cfs
        eb_data["release_af"]  = rel_af
        print(f"  USGS release: {len(rel_cfs.dropna())} records fetched")
    else:
        print(f"  USGS release: no data returned for {EB_RELEASE_USGS_SITE}")

    # Build EB sheets in daily workbook
    ws_eb_storage = wb_daily.create_sheet(title="EB Storage & WSE")
    build_eb_storage_sheet(ws_eb_storage, eb_data)

    ws_eb_release = wb_daily.create_sheet(title="EB Release")
    build_eb_release_sheet(ws_eb_release, eb_data)

    ws_eb_ni = wb_daily.create_sheet(title="EB Net Inflow")
    ni_cfs_series, rel_cfs_series = build_eb_netinflow_sheet(ws_eb_ni, eb_data)

    # Feed EB traces into dashboard using synthetic keys
    dashboard_data["EB_RELEASE"]   = {"flow": rel_cfs_series} if rel_cfs_series is not None else None
    dashboard_data["EB_NETINFLOW"] = {"flow": ni_cfs_series}  if ni_cfs_series  is not None else None
    print("  Done.")

    # --- Save workbooks ---
    daily_file = "RioGrande_WaterYear_Daily.xlsx"
    detail_file = "RioGrande_30Day_15min.xlsx"

    print()  # spacer before save messages
    safe_save(lambda: wb_daily.save(daily_file), daily_file)
    safe_save(lambda: wb_15min.save(detail_file), detail_file)

    # --- Build HTML dashboards ---
    if PLOTLY_AVAILABLE:
        print("\nBuilding daily dashboard...")
        build_dashboard(dashboard_data)
        print("Building 15-minute dashboard...")
        build_15min_dashboard(dashboard_data_15min)

    print("\nAll done. Open the Excel files to view your data.")
    print("Green rows = Approved data   |   Yellow rows = Provisional data")
    if PLOTLY_AVAILABLE:
        print("Open RioGrande_Dashboard.html for the daily water year chart.")
        print("Open RioGrande_Dashboard_15min.html for the 15-minute chart.")


if __name__ == "__main__":
    main()
